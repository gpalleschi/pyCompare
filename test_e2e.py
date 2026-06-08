import sys, os, logging
logging.basicConfig(level=logging.WARNING)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from pycompare.reconcile import ReconcileEngine

# --- Test 1: header-based (source column names) ---
print("=== Test 1: header-based config ===")
engine = ReconcileEngine()
stats = engine.run(
    "test_data/file1.csv", "test_data/file1_config.json",
    "test_data/file2.csv", "test_data/file2_config.json",
    "test_data/match_config.json", "test_data/output",
)
assert stats.total_file1 == 5
assert stats.total_file2 == 4
assert stats.matched == 3, f"Expected 3 matched, got {stats.matched}"
assert stats.only_file1 == 2
assert stats.only_file2 == 1
assert stats.match_percentage == 60.0
print(f"  OK: F1={stats.total_file1} F2={stats.total_file2} M={stats.matched} O1={stats.only_file1} O2={stats.only_file2}")

# --- Test 2: column_index-based config ---
print("=== Test 2: column_index-based config ===")
stats2 = engine.run(
    "test_data/file2.csv", "test_data/file2_config_index.json",
    "test_data/file1.csv", "test_data/file1_config.json",
    "test_data/match_config.json", "test_data/output2",
)
assert stats2.total_file1 == 4
assert stats2.total_file2 == 5
assert stats2.matched == 3, f"Expected 3 matched, got {stats2.matched}"
assert stats2.only_file1 == 1
assert stats2.only_file2 == 2
print(f"  OK: F1={stats2.total_file1} F2={stats2.total_file2} M={stats2.matched} O1={stats2.only_file1} O2={stats2.only_file2}")

# --- Verify output files exist ---
for d in ["output", "output2"]:
    for f in ["reconciled.xlsx", "only_file1.xlsx", "only_file2.xlsx", "statistics.csv"]:
        assert os.path.exists(f"test_data/{d}/{f}"), f"Missing test_data/{d}/{f}"
print("  Output files: OK")

# --- Verify sequence numbers in only_file1.xlsx ---
import openpyxl
wb = openpyxl.load_workbook("test_data/output/only_file1.xlsx")
ws = wb.active
headers = [cell.value for cell in ws[1]]
assert "Row" in headers, f"Expected 'Row' column in only_file1.xlsx, got {headers}"
row_idx = headers.index("Row")
# First data row should have Row=1 (record 4 in file1 has caller=3474567890)
# Records at 0-based indices 3 and 4 → 1-based rows 4 and 5
row_vals = sorted([ws.cell(r, row_idx + 1).value for r in range(2, ws.max_row + 1)])
assert row_vals == [4, 5], f"Expected Row=[4, 5], got {row_vals}"
print(f"  Sequence numbers in only_file1.xlsx: OK (Rows {row_vals})")

wb2 = openpyxl.load_workbook("test_data/output/reconciled.xlsx")
ws2 = wb2.active
headers2 = [cell.value for cell in ws2[1]]
assert "F1_Row" in headers2, f"Expected 'F1_Row' in reconciled.xlsx, got {headers2}"
assert "F2_Row" in headers2, f"Expected 'F2_Row' in reconciled.xlsx, got {headers2}"
print(f"  Sequence numbers in reconciled.xlsx: OK (F1_Row + F2_Row)")

print("\n===== ALL TESTS PASSED =====")
