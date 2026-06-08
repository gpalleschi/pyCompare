from __future__ import annotations
import logging
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from typing import Any

from pycompare.config import FileConfig, MatchConfig

logger = logging.getLogger(__name__)

Record = dict[str, Any]


class ReconciliationStats:
    def __init__(self):
        self.total_file1: int = 0
        self.total_file2: int = 0
        self.matched: int = 0
        self.only_file1: int = 0
        self.only_file2: int = 0
        self.match_percentage: float = 0.0
        self.processing_time: float = 0.0


def generate_reports(
    file1_records: list[Record],
    file2_records: list[Record],
    matched_pairs: list[tuple[int, int]],
    only_file1: list[int],
    only_file2: list[int],
    file1_config: FileConfig,
    file2_config: FileConfig,
    match_config: MatchConfig,
    output_dir: str | Path,
) -> ReconciliationStats:
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    stats = ReconciliationStats()
    stats.total_file1 = len(file1_records)
    stats.total_file2 = len(file2_records)
    stats.matched = len(matched_pairs)
    stats.only_file1 = len(only_file1)
    stats.only_file2 = len(only_file2)
    stats.match_percentage = (stats.matched / stats.total_file1 * 100) if stats.total_file1 > 0 else 0.0

    logger.info("Generating reports in: %s", output_path)

    _write_reconciled(output_path, file1_records, file2_records, matched_pairs, file1_config, file2_config, match_config)
    _write_only(output_path, "only_file1.xlsx", file1_records, only_file1, file1_config)
    _write_only(output_path, "only_file2.xlsx", file2_records, only_file2, file2_config)
    _write_stats_csv(output_path, stats)

    logger.info("Reports generated successfully")
    return stats


def _write_reconciled(
    output_path: Path,
    file1_records: list[Record],
    file2_records: list[Record],
    matched_pairs: list[tuple[int, int]],
    file1_config: FileConfig,
    file2_config: FileConfig,
    match_config: MatchConfig,
):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reconciled"

    f1_field_names = [f.name for f in file1_config.fields]
    f2_field_names = [f.name for f in file2_config.fields]

    header = ["F1_Row", "F2_Row"]
    header.extend([f"F1_{name}" for name in f1_field_names])
    header.extend([f"F2_{name}" for name in f2_field_names])

    diff_fields = _get_diff_fields(match_config, file1_config, file2_config)
    for df in diff_fields:
        header.append(f"Diff_{df}")

    ws.append(header)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for a_idx, b_idx in matched_pairs:
        a_rec = file1_records[a_idx]
        b_rec = file2_records[b_idx]

        row = [a_idx + 1, b_idx + 1]
        for name in f1_field_names:
            row.append(_format_cell(a_rec.get(name)))
        for name in f2_field_names:
            row.append(_format_cell(b_rec.get(name)))
        for df in diff_fields:
            a_val = a_rec.get(df)
            b_val = b_rec.get(df)
            row.append(_compute_diff(a_val, b_val))

        ws.append(row)

    wb.save(str(output_path / "reconciled.xlsx"))
    logger.info("Written: reconciled.xlsx (%d rows)", len(matched_pairs) + 1)


def _write_only(
    output_path: Path,
    filename: str,
    records: list[Record],
    indices: list[int],
    file_config: FileConfig,
):
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    field_names = [f.name for f in file_config.fields]
    ws.append(["Row", *field_names])

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for idx in indices:
        rec = records[idx]
        row = [idx + 1] + [_format_cell(rec.get(name)) for name in field_names]
        ws.append(row)

    wb.save(str(output_path / filename))
    logger.info("Written: %s (%d rows)", filename, len(indices) + 1)


def _get_diff_fields(match_config: MatchConfig, file1_config: FileConfig, file2_config: FileConfig) -> list[str]:
    diff_fields = []
    for mf in match_config.matching_fields:
        fc1 = file1_config.get_field(mf.field)
        if fc1 and fc1.type in ("integer", "decimal", "datetime"):
            diff_fields.append(mf.field)
    return diff_fields


def _compute_diff(a_val: Any, b_val: Any) -> str:
    if a_val is None or b_val is None:
        return ""
    if isinstance(a_val, datetime) and isinstance(b_val, datetime):
        diff_sec = abs((a_val - b_val).total_seconds())
        return f"{diff_sec:.3f}s"
    if isinstance(a_val, (int, float)) and isinstance(b_val, (int, float)):
        return str(abs(a_val - b_val))
    if isinstance(a_val, Decimal) and isinstance(b_val, Decimal):
        return str(abs(a_val - b_val))
    if a_val != b_val:
        return f"{a_val} -> {b_val}"
    return ""


def _format_cell(value: Any) -> Any:
    if isinstance(value, datetime):
        return value
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, int):
        return value
    return str(value) if value is not None else ""


def _write_stats_csv(output_path: Path, stats: ReconciliationStats):
    import csv
    csv_path = output_path / "statistics.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total File 1", stats.total_file1])
        writer.writerow(["Total File 2", stats.total_file2])
        writer.writerow(["Matched", stats.matched])
        writer.writerow(["Only File 1", stats.only_file1])
        writer.writerow(["Only File 2", stats.only_file2])
        writer.writerow(["Match %", f"{stats.match_percentage:.2f}%"])
        writer.writerow(["Processing Time (s)", f"{stats.processing_time:.3f}"])
    logger.info("Written: statistics.csv")
